import tkinter
import tkinter.messagebox
import customtkinter
from tkinter import filedialog
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import excel

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.button_index = 0
        self.upload_buttons = []
        self.uploaded_paths = []

        # configure window
        self.title("CO CALCI")
        self.geometry(f"{1200}x{580}")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=2)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=7, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(7, weight=3)
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="Upload Your Files Here", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=30, pady=(20, 10))

        self.sidebar_button_1 = customtkinter.CTkButton(self.sidebar_frame, text='CAT1',command=self.upload_excel)
        self.sidebar_button_1.grid(row=1, column=0, padx=20, pady=10)
        self.sidebar_button_2 = customtkinter.CTkButton(self.sidebar_frame, text='CAT2' , command=self.upload_excel , state="disabled")
        self.sidebar_button_2.grid(row=2, column=0, padx=20, pady=10)
        self.sidebar_button_3 = customtkinter.CTkButton(self.sidebar_frame,text='CAT3', command=self.upload_excel , state="disabled")
        self.sidebar_button_3.grid(row=3, column=0, padx=20, pady=10)

        self.sidebar_button_4 = customtkinter.CTkButton(self.sidebar_frame,text='ASSIGNMENT' ,command=self.upload_excel , state="disabled")
        self.sidebar_button_4.grid(row=4, column=0, padx=20, pady=10)
        self.sidebar_button_5 = customtkinter.CTkButton(self.sidebar_frame, text='TERMINAL' ,command=self.upload_excel , state="disabled")
        self.sidebar_button_5.grid(row=5, column=0, padx=20, pady=10)
        self.sidebar_button_6 = customtkinter.CTkButton(self.sidebar_frame, text='SURVEY' ,command=self.upload_excel , state="disabled")
        self.sidebar_button_6.grid(row=6, column=0, padx=20, pady=10)

        self.sidebar_button_7 = customtkinter.CTkButton(self.sidebar_frame, text='CALCULATE' ,command=self.compute_ans , state="disabled")
        self.sidebar_button_7.grid(row=7, column=0, padx=20, pady=10)

        self.upload_buttons=[self.sidebar_button_1,self.sidebar_button_2,self.sidebar_button_3,self.sidebar_button_4,self.sidebar_button_5,self.sidebar_button_6,self.sidebar_button_7]
        
        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=8, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=9, column=0, padx=20, pady=(10, 10))
        self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        self.scaling_label.grid(row=10, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
                                                               command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=10, column=0, padx=20, pady=(10, 20))


        self.centerbar_frame = customtkinter.CTkFrame(self, width=130, corner_radius=0)
        self.centerbar_frame.grid(row=0, column=1, rowspan=8, sticky="nsew")
        self.centerbar_frame.grid_rowconfigure(8, weight=1)
        
        self.input_label = customtkinter.CTkLabel(self.centerbar_frame, text="Give Your Inputs Here", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.input_label.grid(row=1, column=1, padx=30, pady=20)

        self.input_label = customtkinter.CTkLabel(self.centerbar_frame, text="Percentage Input", font=customtkinter.CTkFont(size=16, weight="normal"))
        self.input_label.grid(row=3, column=1, padx=30, pady=10)

        self.perc_input = customtkinter.CTkEntry(self.centerbar_frame, width=220,placeholder_text="Give Float Input Here",state="disable")
        self.perc_input.grid(row=4, column=1,columnspan=1 ,padx=50, pady=10, sticky="nsew")

        self.input_label = customtkinter.CTkLabel(self.centerbar_frame, text="Attainmemt Input", font=customtkinter.CTkFont(size=16, weight="normal"))
        self.input_label.grid(row=5, column=1, padx=30, pady=10)

        self.att_input = customtkinter.CTkEntry(self.centerbar_frame, width=220,placeholder_text="Give Float Input Here",state= "disable")
        self.att_input.grid(row=6, column=1,columnspan=1 ,padx=50, pady=10, sticky="nsew")


        self.main_button_1 = customtkinter.CTkButton(master=self.centerbar_frame,command=self.get_user_input, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"))
        self.main_button_1.grid(row=7, column=1, padx=30, pady=10, sticky="nsew")


        self.print_area = customtkinter.CTkLabel(self,text="console")
        self.print_area.grid(row=2,column=1,columnspan=2,padx=20,pady=5)

        self.tabview = customtkinter.CTkTabview(self, width=250)
        self.tabview.grid(row=0, column=2,columnspan =2, padx=(20, 0), pady=(20, 0), sticky="nsew")
        self.tabview.add("CO-SUMMARY")
        self.tabview.tab("CO-SUMMARY").grid_columnconfigure(0, weight=1)


        # create scrollable frame
        self.scrollable_frame = customtkinter.CTkScrollableFrame(self, label_text="CTkScrollableFrame")
        self.scrollable_frame.grid(row=1, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")
        self.scrollable_frame.grid_columnconfigure(0, weight=1)
        self.scrollable_frame_switches = []
        # for i in range(100):
        #     switch = customtkinter.CTkSwitch(master=self.scrollable_frame, text=f"CTkSwitch {i}")
        #     switch.grid(row=i, column=0, padx=10, pady=(0, 20))
        #     self.scrollable_frame_switches.append(switch)

        # create checkbox and switch frame
        self.checkbox_slider_frame = customtkinter.CTkFrame(self)
        self.checkbox_slider_frame.grid(row=1, column=3, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.heading_lable = customtkinter.CTkLabel(self.checkbox_slider_frame,text='Download Your Files Here',font=customtkinter.CTkFont(size=20, weight="bold"))
        self.heading_lable.grid(row=0, column=0, padx=30, pady=(20, 10))
        

        # set default values
        # self.checkbox_3.configure(state="disabled")
        # self.checkbox_1.select()
        # self.scrollable_frame_switches[0].select()
        # self.scrollable_frame_switches[4].select()
        # self.appearance_mode_optionemenu.set("Dark")
        # self.scaling_optionemenu.set("100%")

    def upload_excel(self):
        if self.button_index <= len(self.upload_buttons) - 1:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                print(f"Uploaded file for button {self.button_index + 1}: {file_path}")
                self.print_area.configure(text=f"Uploaded file for button {self.button_index + 1}: {file_path}")
                self.uploaded_paths.append(file_path)
                self.upload_buttons[self.button_index].configure(state="disabled")
                self.button_index += 1
                self.upload_buttons[self.button_index].configure(state="normal")

    def compute_ans(self):

        for i in range(0,3):
            test_book = load_workbook(self.uploaded_paths[i])
            test_sheet = test_book.active
            excel.foreveryrow(test_sheet,excel.find_max_rows(test_sheet))
        excel.calc_without_assignment()

        assignment_book = load_workbook(self.uploaded_paths[3])
        assignment_sheet = assignment_book.active
        excel.compute_assignment(assignment_sheet)

        terminal_book = load_workbook(self.uploaded_paths[4])
        terminal_sheet = terminal_book.active

        survey_book = load_workbook(self.uploaded_paths[5])
        survey_sheet = survey_book.active

        excel.add_survey_terminal(survey_sheet,terminal_sheet)
        self.make_available()

        return

    def open_input_dialog_event(self):
        dialog = customtkinter.CTkInputDialog(text="Type in a number:", title="CTkInputDialog")
        print("CTkInputDialog:", dialog.get_input())

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)

    def sidebar_button_event(self):
        print("sidebar_button click")

    def upload_assignment_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            test_object = load_workbook(file_path)
            test_object=test_object.active
            excel.compute_assignment()
    
    def make_every_tabs(self,col_list,no_students,actual_att,tot_att):
        self.tabview.add(col_list)
        self.tabview.tab(col_list).grid_columnconfigure(0, weight=1)
        self.total_students = customtkinter.CTkLabel(self.tabview.tab(col_list),text=f'Total Number Of Students Above Percentage : {no_students}')
        self.total_students.grid(row=1, column=0, padx=10, pady=10)
        self.Actual_attainment = customtkinter.CTkLabel(self.tabview.tab(col_list),text=f'Actual Attainment of students : {actual_att}')
        self.Actual_attainment.grid(row=2, column=0, padx=10, pady=10)
        self.Actual_attainment = customtkinter.CTkLabel(self.tabview.tab(col_list),text=f'Total Attainment of students : {tot_att}')
        self.Actual_attainment.grid(row=3, column=0, padx=10, pady=10)      
        return
    
    def download(self,source):

        # file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
        # if file_path:
        #     source = "summary.xlsx"
        # try:
        #     with open(source, 'rb') as source, open(file_path, 'wb') as target:
        #         target.write(source.read())
        #     print(f"'{source}' saved to '{file_path}'")
        # except FileNotFoundError:
        #     print(f"'{source}' not found.")
        return
    
    def make_available(self):
        self.perc_input.configure(state="normal")
        self.att_input.configure(state="normal")
        return
        

    def make_download(self):
        self.download_button = customtkinter.CTkButton(self.checkbox_slider_frame, text="Internals_summary", command = self.download('new_summary_data.xlsx'))
        self.download_button.grid(row=2, column=0, padx=20, pady=10)

        self.download_button = customtkinter.CTkButton(self.checkbox_slider_frame, text="Overall_summary", command = self.download('final_summary.xlsx'))
        self.download_button.grid(row=3, column=0, padx=20, pady=10)

    def give_summary(self,total_summary):
        co_list = total_summary[0]
        no_students = total_summary[1]
        actual_att = total_summary[2]
        tot_att = total_summary[3]

        for cos in co_list:
            self.make_every_tabs(cos,no_students[cos],actual_att[cos],tot_att[cos])
        return

    def get_user_input(self):
        percentage_input = float(self.perc_input.get())
        attainment_input = float(self.att_input.get())
        total_summary = excel.calc_final_percentage(percentage_input,attainment_input)

        self.total_number = customtkinter.CTkLabel(self.tabview.tab("CO-SUMMARY"),text=f'Total No of Students : {excel.total_students}')
        self.total_number.grid(row=1, column=0, padx=10, pady=10)
        
        self.percentage = customtkinter.CTkLabel(self.tabview.tab("CO-SUMMARY"),text=f'Given Percentage : {percentage_input}')
        self.percentage.grid(row=2, column=0, padx=10, pady=10)

        self.Attainment = customtkinter.CTkLabel(self.tabview.tab("CO-SUMMARY"),text=f'Given Attainment : {attainment_input}')
        self.Attainment.grid(row=3, column=0, padx=10, pady=10)
        self.give_summary(total_summary)
        self.make_download()

    

if __name__ == "__main__":
    app = App()
    app.resizable(False,False)
    app.mainloop()