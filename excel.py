import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
# wb = load_workbook('14CS350_Data Structures and Algorithms.xlsx')

total_students = 0 
no_co=0
co_list=[]
 

# test1 =load_workbook('CAT1.xlsx').active
# print(test1)

# test2 = load_workbook('CAT2.xlsx').active
# print(test2)

# test3 =load_workbook('CAT3.xlsx').active
# print(test3)



def find_max_rows(test):
    max_row_with_data =0
    for row in test.iter_rows(values_only=True):
        if any(cell is not None  for cell in row):
            max_row_with_data += 1
    return max_row_with_data

def find_no_columns(test,row):
    c1=1
    for col in test[row]:
        if col!=None:
            c1=c1+1
        else:
            break
    return c1

def create_workbook(title,name):
    Wb_test = Workbook()
    sheet = Wb_test.active
    sheet.title = title
    Wb_test.save(name)



create_workbook('overall','new_summary_data.xlsx')

def map_part(test):
    for col in range(1,test.max_col):
        print()



def create_tot(test,row):
    i=1
    dict={}
    row = test[3][1:]

    for cos in row:
        if cos.value ==  None:
            return dict
        if cos.value not in dict:
            dict[cos.value]=0
        dict[cos.value]+= int(test[4][i].value) if test[4][i].value != None else 0
        i=i+1
    return dict

def create_sum(test,r1):
    dict={}
    i=1
    row = test[3][1:]
    for cos in row:
        if cos.value ==  None:
            return dict
        if cos.value not in dict:
            dict[cos.value]=0
        dict[cos.value]+= int(test[r1][i].value) if test[r1][i].value != None else 0
        i=i+1
    return dict

def create_tot_sum(test,row_given):
    dict = {}
    i=1
    row = test[3][1:]
    for cos in range(0,len(row)+1):
        val = row[cos].value
        if val == None:
            return dict
        if val not in dict :
            dict[val]=0  
        if test[row_given][cos+1].value is not None and test[row_given][cos+1].value >=0:
            dict[val] += test[4][cos+1].value
    return dict


def foreveryrow(test,max_row):
    total_dict = create_tot(test,1)
    temp_list = list(total_dict.keys())
    choice_co = temp_list[-1]
    total_dict[choice_co] = total_dict[choice_co]/2
    
    new_sheet = load_workbook('new_summary_data.xlsx')
    new_wb = new_sheet['overall']
    # new_wb.insert_rows(test.max_row)
    new_sheet.save('new_summary_data.xlsx')
    i=0
    for regno in range(5,test.max_row):
        new_wb.cell(row= regno-2, column= 1, value=test[regno][0].value)
    new_sheet.save('new_summary_data.xlsx')

    c2=1
    for col in new_wb[3]:
        if col.value!=None:
            c2=c2+1
        else :
            break
    c=c2

    for keys in total_dict:
        new_wb.cell(row=1,column=c,value=keys)
        new_wb.cell(row=2,column=c,value=total_dict[keys])
        new_wb.cell(row=1,column=c+len(total_dict),value= str(keys)+'%')
        c= c + 1
    new_sheet.save('new_summary_data.xlsx')

    c1=1
    for col in new_wb[3]:
        if col.value!=None:
            c1=c1+1
        else :
            break

    for rows in range(5,max_row):
        val_dict = create_sum(test,rows)
        eval_dict = create_tot_sum(test,rows)
        c=c1
        for keys in val_dict:
            new_wb.cell(row=rows-2,column=c,value=val_dict[keys])  
            if eval_dict[keys] ==0:
                new_wb.cell(row=rows-2,column= c + len(val_dict),value =0)
            else:
                new_wb.cell(row=rows-2,column=c+len(val_dict),value=round(val_dict[keys]/eval_dict[keys] ,2)*100) 
            c=c+1
    new_sheet.save('new_summary_data.xlsx')
    return new_sheet

# def calculate_overall():
    
#     for each_test in test_list:
#         foreveryrow(each_test,max_row)

def find_cos(wb,row):
    new_dict={}
    for col in wb[1]:
        co = str(col.value)
        if co.endswith('%'):
            co =co[:-1]
            if co not in new_dict:
                new_dict[co]=[]
            new_dict[co].append(wb[row][col.column-1].value)
    return new_dict


def find_max_pair(percentage_list):
    
    max_sum=0
    
    if(len(percentage_list)==1):
        max_sum = percentage_list[0]
    if(len(percentage_list)==2):
        max_sum = round((percentage_list[0]+percentage_list[1])/2,2)
    if(len(percentage_list)==3):
        for i in range(0,len(percentage_list)):
            for j in range(i+1,len(percentage_list)):
                max_sum = max(round((percentage_list[i] + percentage_list[j])/2,2),max_sum)
    return (max_sum + calc_assignment())/2


def calc_assignment():
    return 100

def compute_assignment(test):
    return 0

def calc_without_assignment():

    # calculate_overall()
    new_sheet = load_workbook('new_summary_data.xlsx')
    new_wb = new_sheet['overall']

    create_workbook('final','final_summary.xlsx')
    new_final = load_workbook('final_summary.xlsx')
    new_fl = new_final['final']
    new_fl.cell(row=1,column=1,value='Internals')

    col_val = find_no_columns(new_wb,1)
    for rows in range(3,new_wb.max_row):
        percentage_dict = find_cos(new_wb,rows)
        i=1

        for keys in percentage_dict.keys():
            if rows==3:
                new_fl.cell(row=2,column=i,value=str(keys))
                global co_list
                co_list = list(percentage_dict.keys())
                co_list.sort()
            new_percentage_list= percentage_dict[keys]
            max_val= find_max_pair(new_percentage_list)
            new_fl.cell(row=rows,column=i,value=max_val)
            i=i+1
    data = list(new_fl.iter_rows(values_only=True))
    data_transposed = list(zip(*data))
    sorted_data_transposed = sorted(data_transposed, key=lambda x: x[2 - 1])
    sorted_data = list(zip(*sorted_data_transposed))
    new_fl.delete_rows(1, new_fl.max_row)
    for row in sorted_data:
        new_fl.append(row)

    
    new_sheet.save('new_summary_data.xlsx')
    new_final.save('final_summary.xlsx')

# def process_assignment():


def add_survey_terminal(survey_sheet,terminal_sheet):

    global total_students 
    total_students = find_max_rows(terminal_sheet) -2

    new_final = load_workbook('final_summary.xlsx')
    new_fl=new_final['final']

    col_val = find_no_columns(new_fl,1)
    new_fl.cell(row=1,column= col_val,value='Terminal')

    start_row=0
    start_column= col_val -1

    for row in terminal_sheet.iter_rows(min_row=2):
        for cell in row:
            new_fl.cell(
                row= start_row + cell.row,
                column=start_column + cell.column,
                value=cell.value
            )

    col_val =find_no_columns(new_fl,1)
    new_fl.cell(row=1,column=col_val,value='Survey')

    print(survey_sheet.max_row)
    start_row=0
    start_column = col_val -1 

    for row in survey_sheet.iter_rows(min_row=2):
        for cell in row:
            new_fl.cell(
                row= start_row + cell.row,
                column=start_column + cell.column,
                value= cell.value if cell.value else 0
            )

    col_val = find_no_columns(new_fl,1)
    new_fl.cell(row=1,column=col_val,value='Attainment')

    start_row =0
    global no_co
    no_co = survey_sheet.max_column
    start_column = col_val - 1
    for row in range(3,survey_sheet.max_row + 1):
            for i in range(0,no_co ):
                val1=new_fl[row][i].value 
                val2 = new_fl[row][i + no_co].value
                val3 = new_fl[row][i + 2*no_co ].value

                val1 = val1 if val1 else 0
                val2 = val2 if val2 else 0
                val3 = val3 if val3 else 0
                final_val = (val1*0.60) + ( val2* 0.30) + ( val3 * 0.10)
                new_fl.cell(row = row,column = start_column + i + 1 , value= final_val)

                if row ==3:
                    new_fl.cell(row = 2,column= start_column + i +1 ,value= new_fl[2][i].value)

    new_final.save('final_summary.xlsx')

def count_particaular_range(perc,co,test):
    count=0
    for rows in test.iter_rows(min_row = 3 ,values_only = True):
        if float(rows[co]) >= perc :
            count = count+1
    return count

def calc_final_percentage(given_percentage,attainment):
    final_summary_dict ={}
    no_students={}
    actual_att={}
    outcome_att={}
    new_final = load_workbook('final_summary.xlsx')
    new_fl = new_final.active
    first_co = 3*no_co
    print(co_list)
    for i in range(0,no_co):
        final_summary_dict[co_list[i]] =[]
        no_students[co_list[i]]=count_particaular_range(given_percentage,first_co+i,new_fl)
        actual_att[co_list[i]]= round(no_students[co_list[i]]/total_students,3)
        outcome_att[co_list[i]] = round(actual_att[co_list[i]]/attainment,3)
        print(f'{co_list[i]} : actual attainment : {actual_att[co_list[i]]}  outcome attainment : {outcome_att[co_list[i]]}')
    return [co_list,no_students,actual_att,outcome_att]
# calc_without_assignment()
# add_survey_terminal()
# calc_final_percentage()